import calendar
from datetime import datetime, timedelta, date
import pandas as pd
import holidays
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

FEIERTAGSZUSCHLAG = 1.5

def get_saxony_holidays(jahr, monat):
    de_holidays = holidays.DE(prov='SN', years=jahr)
    feiertage = []
    for datum, name in de_holidays.items():
        if datum.month == monat:
            feiertage.append(datum.day)
    return feiertage

def apply_borders(worksheet, start_row, end_row, start_col, end_col):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = thin_border

def create_month_sheet(worksheet, jahr, monat, monatsnamen, use_holiday_bonus, stundenlohn):
    _, tage_im_monat = calendar.monthrange(jahr, monat)
    feiertage = get_saxony_holidays(jahr, monat)
    
    # Erstelle Zeitintervalle (Stunden)
    zeitintervalle = []
    for stunde in range(24):
        ende = "24" if stunde == 23 else f"{(stunde + 1):02d}"
        zeitintervalle.append(f"{stunde:02d}-{ende}")
    
    # Kopfzeile
    worksheet.cell(row=1, column=1, value=f"{monatsnamen[monat]} {jahr}")
    
    # Spaltenüberschriften (Tage)
    worksheet.cell(row=4, column=1, value="Stunden")
    for tag in range(1, tage_im_monat + 1):
        worksheet.cell(row=4, column=tag + 1, value=f"{tag:02d}")
    # Zusätzliche Spalte für Monatssumme
    worksheet.cell(row=4, column=tage_im_monat + 2, value="Gesamt")
    
    # Zeitintervalle
    for idx, zeit in enumerate(zeitintervalle, start=5):
        worksheet.cell(row=idx, column=1, value=zeit)
    
    # Gehaltszeilen
    row_offset = len(zeitintervalle) + 4
    worksheet.cell(row=row_offset + 1, column=1, value="Stunden gesamt")
    worksheet.cell(row=row_offset + 2, column=1, value="Grundlohn")
    if use_holiday_bonus:
        worksheet.cell(row=row_offset + 3, column=1, value="Feiertagszuschlag")
    else:
        worksheet.cell(row=row_offset + 3, column=1, value="")  # Leere Zeile für Feiertagszuschlag
    worksheet.cell(row=row_offset + 4, column=1, value="Gesamtlohn")
    
    # Styles
    gelb_fuellen = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_fuellen = PatternFill(start_color='FFD8B2', end_color='FFD8B2', fill_type='solid')
    
    # Färbe Stunden-Zeile orange
    for col in range(1, tage_im_monat + 3):  # +3 für die erste und die Gesamtspalte
        worksheet.cell(row=4, column=col).fill = orange_fuellen
    
    # Setze Formeln und Formatierung für jeden Tag
    for tag in range(1, tage_im_monat + 1):
        spalte = get_column_letter(tag + 1)
        
        # Markiere Feiertage gelb (bleibt für beide Varianten gleich)
        if tag in feiertage:
            for row in range(5, 5 + len(zeitintervalle)):
                worksheet.cell(row=row, column=tag + 1).fill = gelb_fuellen
        
        # Formeln
        # Stunden gesamt
        stunden_formel = f'=SUM({spalte}5:{spalte}28)'
        worksheet.cell(row=row_offset + 1, column=tag + 1, value=stunden_formel)
        
        # Grundlohn
        worksheet.cell(row=row_offset + 2, column=tag + 1, value=f"={spalte}{row_offset + 1}*{stundenlohn}")
        
        if use_holiday_bonus:
            feiertagsformel = f"={spalte}{row_offset + 2}*0.5" if tag in feiertage else "=0"
            worksheet.cell(row=row_offset + 3, column=tag + 1, value=feiertagsformel)
            worksheet.cell(row=row_offset + 4, column=tag + 1, value=f"={spalte}{row_offset + 2}+{spalte}{row_offset + 3}")
        else:
            worksheet.cell(row=row_offset + 3, column=tag + 1, value="")  # Leere Zelle für Feiertagszuschlag
            worksheet.cell(row=row_offset + 4, column=tag + 1, value=f"={spalte}{row_offset + 2}")  # Nur Grundlohn
    
    # Gesamtspalte Formeln
    gesamt_spalte = get_column_letter(tage_im_monat + 2)
    
    # Stunden gesamt für den Monat
    worksheet.cell(row=row_offset + 1, column=tage_im_monat + 2, 
                  value=f"=SUM(B{row_offset + 1}:{get_column_letter(tage_im_monat + 1)}{row_offset + 1})")
    
    # Grundlohn gesamt für den Monat
    worksheet.cell(row=row_offset + 2, column=tage_im_monat + 2, 
                  value=f"=SUM(B{row_offset + 2}:{get_column_letter(tage_im_monat + 1)}{row_offset + 2})")
    
    if use_holiday_bonus:
        # Feiertagszuschlag gesamt für den Monat
        worksheet.cell(row=row_offset + 3, column=tage_im_monat + 2, 
                      value=f"=SUM(B{row_offset + 3}:{get_column_letter(tage_im_monat + 1)}{row_offset + 3})")
        
        # Gesamtlohn für den Monat (mit Zuschlag)
        worksheet.cell(row=row_offset + 4, column=tage_im_monat + 2, 
                      value=f"=SUM(B{row_offset + 4}:{get_column_letter(tage_im_monat + 1)}{row_offset + 4})")
    else:
        # Leere Zelle für Feiertagszuschlag
        worksheet.cell(row=row_offset + 3, column=tage_im_monat + 2, value="")
        
        # Gesamtlohn für den Monat (ohne Zuschlag)
        worksheet.cell(row=row_offset + 4, column=tage_im_monat + 2, 
                      value=f"=SUM(B{row_offset + 4}:{get_column_letter(tage_im_monat + 1)}{row_offset + 4})")

    # Füge Gitterlinien hinzu
    apply_borders(worksheet, 4, 28, 1, tage_im_monat + 2)  # +2 für die Gesamtspalte
    
    # Entferne die Rahmenlinien für die letzten 4 Zeilen (Berechnungen)
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    
    for row in range(29, 33):  # Zeilen für Berechnungen
        for col in range(1, tage_im_monat + 3):  # +3 für die erste und die Gesamtspalte
            worksheet.cell(row=row, column=col).border = no_border

def create_overview_sheet(workbook, monatsnamen):
    overview = workbook.create_sheet("Übersicht", 0)
    overview.cell(row=1, column=1, value="Jahresübersicht")
    
    # Überschriften
    overview.cell(row=3, column=1, value="Monat")
    overview.cell(row=3, column=2, value="Gesamtstunden")
    overview.cell(row=3, column=3, value="Gesamtlohn")
    
    # Daten für jeden Monat
    for monat in range(1, 13):
        overview.cell(row=monat + 3, column=1, value=monatsnamen[monat])
        # Gesamtstunden - Summe der Stunden des jeweiligen Monats
        overview.cell(row=monat + 3, column=2, 
                     value=f'=SUM(\'{monatsnamen[monat]}\'!B29:AE29)')
        # Gesamtlohn - Summe des Gesamtlohns des jeweiligen Monats
        overview.cell(row=monat + 3, column=3, 
                     value=f'=SUM(\'{monatsnamen[monat]}\'!B32:AE32)')
    
    # Jahresgesamtwerte
    overview.cell(row=16, column=1, value="Jahresgesamt")
    # Gesamtstunden des Jahres
    overview.cell(row=16, column=2, value="=SUM(B4:B15)")
    # Gesamtlohn des Jahres
    overview.cell(row=16, column=3, value="=SUM(C4:C15)")
    
    # Füge Gitterlinien zur Übersicht hinzu
    apply_borders(overview, 3, 16, 1, 3)

def create_monthly_schedule(dateiname=None, use_holiday_bonus=False, stundenlohn=12.50):
    heute = datetime.now()
    jahr = heute.year
    
    monatsnamen = {
        1: "Januar", 2: "Februar", 3: "März", 4: "April",
        5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
        9: "September", 10: "Oktober", 11: "November", 12: "Dezember"
    }
    
    # Bestimme den Dateinamen
    if dateiname:
        excel_dateiname = f"Zeitplan_{dateiname}_{jahr}.xlsx"
    else:
        excel_dateiname = f"Zeitplan_{jahr}.xlsx"
    
    # Erstelle Excel-Workbook
    workbook = Workbook()
    
    # Erstelle Übersichtssheet
    create_overview_sheet(workbook, monatsnamen)
    
    # Erstelle Monatssheets
    for monat in range(1, 13):
        sheet = workbook.create_sheet(monatsnamen[monat])
        create_month_sheet(sheet, jahr, monat, monatsnamen, use_holiday_bonus, stundenlohn)
    
    # Entferne das standardmäßig erstellte Sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    
    # Speichere Workbook
    workbook.save(excel_dateiname)
    return excel_dateiname