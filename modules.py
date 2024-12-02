import calendar
from datetime import datetime, timedelta, date
import pandas as pd
import holidays
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

FEIERTAGSZUSCHLAG = 1.5


def get_saxony_holidays(bundesland, jahr, monat):
    bl = bundesland
    de_holidays = holidays.DE(prov=bl, years=jahr)
    feiertage = []
    for datum, name in de_holidays.items():
        if datum.month == monat:
            feiertage.append(datum.day)
    return feiertage

def apply_borders(worksheet, start_row, end_row, start_col, end_col):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = thin_border

def create_month_sheet(worksheet, jahr, monat, monatsnamen, bundesland, use_holiday_bonus, stundenlohn):
    _, tage_im_monat = calendar.monthrange(jahr, monat)
    feiertage = get_saxony_holidays(bundesland, jahr, monat)
    
    # Erstelle Zeitintervalle (Stunden)
    zeitintervalle = []
    for stunde in range(24):
        ende = "24" if stunde == 23 else f"{(stunde + 1):02d}"
        zeitintervalle.append(f"{stunde:02d}-{ende}")
    
    # Kopfzeile
    worksheet.cell(row=1, column=1, value=f"{monatsnamen[monat]} {jahr}")
    
    # Spaltenüberschriften (Tage)
    worksheet.cell(row=4, column=1, value="Stunden")
    for tag in range(1, tage_im_monat + 1):
        worksheet.cell(row=4, column=tag + 1, value=f"{tag:02d}")
    worksheet.cell(row=4, column=tage_im_monat + 2, value="Gesamt")
    
    # Zeitintervalle
    for idx, zeit in enumerate(zeitintervalle, start=5):
        worksheet.cell(row=idx, column=1, value=zeit)
    
    # Gehaltszeilen
    row_offset = len(zeitintervalle) + 4
    worksheet.cell(row=row_offset + 1, column=1, value="Stunden gesamt")
    worksheet.cell(row=row_offset + 2, column=1, value="Grundlohn")
    if use_holiday_bonus:
        worksheet.cell(row=row_offset + 3, column=1, value="Feiertagszuschlag")
    else:
        worksheet.cell(row=row_offset + 3, column=1, value="")
    worksheet.cell(row=row_offset + 4, column=1, value="Gesamtlohn")
    
    # Styles
    gelb_fuellen = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_fuellen = PatternFill(start_color='FFD8B2', end_color='FFD8B2', fill_type='solid')
    
    # Färbe Stunden-Zeile orange
    for col in range(1, tage_im_monat + 3):
        worksheet.cell(row=4, column=col).fill = orange_fuellen
    
    # Setze Formeln und Formatierung für jeden Tag
    for tag in range(1, tage_im_monat + 1):
        spalte = get_column_letter(tag + 1)
        
        # Markiere Feiertage gelb
        if tag in feiertage:
            for row in range(5, 5 + len(zeitintervalle)):
                worksheet.cell(row=row, column=tag + 1).fill = gelb_fuellen
        
        # Formeln mit ROUND-Funktion für 2 Dezimalstellen
        # Stunden gesamt
        stunden_formel = f'=ROUND(SUM({spalte}5:{spalte}28), 2)'
        worksheet.cell(row=row_offset + 1, column=tag + 1, value=stunden_formel)
        
        # Grundlohn
        #worksheet.cell(row=row_offset + 2, column=tag + 1, value=f"=ROUND({spalte}{row_offset + 1}*{stundenlohn}, 2)")
        worksheet.cell(row=row_offset + 2, column=tag + 1, value=f"=ROUND({spalte}{row_offset + 1}*\'Übersicht\'!E4, 2)")
        
        if use_holiday_bonus:
            feiertagsformel = f"=ROUND({spalte}{row_offset + 2}*0.5, 2)" if tag in feiertage else "=0"
            worksheet.cell(row=row_offset + 3, column=tag + 1, value=feiertagsformel)
            worksheet.cell(row=row_offset + 4, column=tag + 1, value=f"=ROUND({spalte}{row_offset + 2}+{spalte}{row_offset + 3}, 2)")
        else:
            worksheet.cell(row=row_offset + 3, column=tag + 1, value="")
            worksheet.cell(row=row_offset + 4, column=tag + 1, value=f"=ROUND({spalte}{row_offset + 2}, 2)")
    
    # Gesamtspalte Formeln
    gesamt_spalte = get_column_letter(tage_im_monat + 2)
    
    # Stunden gesamt für den Monat
    worksheet.cell(row=row_offset + 1, column=tage_im_monat + 2, 
                  value=f"=ROUND(SUM(B{row_offset + 1}:{get_column_letter(tage_im_monat + 1)}{row_offset + 1}), 2)")
    
    # Grundlohn gesamt für den Monat
    worksheet.cell(row=row_offset + 2, column=tage_im_monat + 2, 
                  value=f"=ROUND(SUM(B{row_offset + 2}:{get_column_letter(tage_im_monat + 1)}{row_offset + 2}), 2)")
    
    if use_holiday_bonus:
        # Feiertagszuschlag gesamt für den Monat
        worksheet.cell(row=row_offset + 3, column=tage_im_monat + 2, 
                      value=f"=ROUND(SUM(B{row_offset + 3}:{get_column_letter(tage_im_monat + 1)}{row_offset + 3}), 2)")
        
        # Gesamtlohn für den Monat (mit Zuschlag)
        worksheet.cell(row=row_offset + 4, column=tage_im_monat + 2, 
                      value=f"=ROUND(SUM(B{row_offset + 4}:{get_column_letter(tage_im_monat + 1)}{row_offset + 4}), 2)")
    else:
        worksheet.cell(row=row_offset + 3, column=tage_im_monat + 2, value="")
        worksheet.cell(row=row_offset + 4, column=tage_im_monat + 2, 
                      value=f"=ROUND(SUM(B{row_offset + 4}:{get_column_letter(tage_im_monat + 1)}{row_offset + 4}), 2)")

    # Füge Gitterlinien hinzu
    apply_borders(worksheet, 4, 28, 1, tage_im_monat + 2)
    
    # Entferne die Rahmenlinien für die letzten 4 Zeilen
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    
    for row in range(29, 33):
        for col in range(1, tage_im_monat + 3):
            worksheet.cell(row=row, column=col).border = no_border

def create_overview_sheet(workbook, monatsnamen, stundenlohn):
    overview = workbook.create_sheet("Übersicht", 0)
    overview.cell(row=1, column=1, value="Jahresübersicht")
    
    # Überschriften
    overview.cell(row=3, column=1, value="Monat")
    overview.cell(row=3, column=2, value="Gesamtstunden")
    overview.cell(row=3, column=3, value="Gesamtlohn")
    overview.cell(row=3, column=5, value="Stundenlohn")
    overview.cell(row=4, column=5, value=stundenlohn)
    
    # Daten für jeden Monat mit ROUND-Funktion
    for monat in range(1, 13):
        overview.cell(row=monat + 3, column=1, value=monatsnamen[monat])
        overview.cell(row=monat + 3, column=2, 
                     value=f'=ROUND(SUM(\'{monatsnamen[monat]}\'!B29:AE29), 2)')
        overview.cell(row=monat + 3, column=3, 
                     value=f'=ROUND(SUM(\'{monatsnamen[monat]}\'!B32:AE32), 2)')
    
    # Jahresgesamtwerte mit ROUND-Funktion
    overview.cell(row=16, column=1, value="Jahresgesamt")
    overview.cell(row=16, column=2, value="=ROUND(SUM(B4:B15), 2)")
    overview.cell(row=16, column=3, value="=ROUND(SUM(C4:C15), 2)")
    
    # Füge Gitterlinien zur Übersicht hinzu
    apply_borders(overview, 3, 16, 1, 3)

def create_monthly_schedule(dateiname=None, bundesland=None, stundenlohn=None, use_holiday_bonus=False):
    heute = datetime.now()
    jahr = heute.year
    monatsnamen = {
        1: "Januar", 2: "Februar", 3: "März", 4: "April",
        5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
        9: "September", 10: "Oktober", 11: "November", 12: "Dezember"
    }
    
    if dateiname:
        excel_dateiname = f"Zeitplan_{dateiname}_{jahr}.xlsx"
    else:
        excel_dateiname = f"Zeitplan_{jahr}.xlsx"
    
    workbook = Workbook()
    create_overview_sheet(workbook, monatsnamen, stundenlohn)
    
    for monat in range(1, 13):
        sheet = workbook.create_sheet(monatsnamen[monat])
        create_month_sheet(sheet, jahr, monat, monatsnamen, bundesland, use_holiday_bonus, stundenlohn)
    
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    
    workbook.save(excel_dateiname)
    return excel_dateiname